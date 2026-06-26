# 売上利益ブックを業務用デザインに一括整形するスタイラー
# 既存データは変更せず、見た目（配色・行縞・桁区切り・列幅・固定）だけを整える。
import re, sys, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule

NAVY   = "1F3A5F"
NAVY_D = "16293F"
ACCENT = "C8A24B"
BAND   = "EEF3FA"
TOTAL  = "DCE5F4"   # 合計行の地色（淡い紺）
WHITE  = "FFFFFF"
GRID   = "D7DEE8"
TXT    = "1A2433"
GREEN  = "1E7B45"
AMBER  = "B5740F"
RED    = "C0392B"   # マイナス数値

thin = Side(style="thin", color=GRID)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
TOP_NAVY = Side(style="medium", color=NAVY)

TOTAL_RE = re.compile(r"(合計|総計|小計|計$|期計|累計計|全体)")
def is_total_row(ws, r, hdr_r, ncol):
    for c in range(1, min(ncol, 3) + 1):
        v = ws.cell(r, c).value
        if v not in (None, "") and TOTAL_RE.search(str(v)):
            return True
    return False

def cjk_w(s):
    s = "" if s is None else str(s)
    return sum(2 if ord(c) > 0x2E7F else 1 for c in s)

def is_num_col(h):
    h = str(h or "")
    keys = ["万", "件数", "率", "比", "順位", "価格", "金額", "累計", "合計", "売上", "利益", "粗利", "歩合", "原価", "値"]
    return any(k in h for k in keys)

def is_pct_col(h):
    h = str(h or "")
    return ("率" in h) or ("比" in h)

def find_rows(ws):
    title_r = None; hdr_r = None
    for r in range(1, min(ws.max_row, 14) + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        ne = [v for v in vals if v not in (None, "")]
        if not ne:
            continue
        if title_r is None and len(ne) == 1:
            title_r = r; continue
        if hdr_r is None and len(ne) >= 2:
            hdr_r = r; break
    return title_r, hdr_r

def beautify_sheet(ws):
    title_r, hdr_r = find_rows(ws)
    if hdr_r is None:
        return
    ncol = ws.max_column
    nrow = ws.max_row
    ws.sheet_view.showGridLines = False
    from openpyxl.formatting.formatting import ConditionalFormattingList
    ws.conditional_formatting = ConditionalFormattingList()  # 再実行で重複しないよう初期化

    if title_r is not None:
        tc = ws.cell(title_r, 1)
        tc.font = Font(name="Yu Gothic", size=15, bold=True, color=NAVY_D)
        tc.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[title_r].height = 30
        try:
            ws.merge_cells(start_row=title_r, start_column=1, end_row=title_r, end_column=ncol)
        except Exception:
            pass
        tc.font = Font(name="Yu Gothic", size=14, bold=True, color=WHITE)
        for c in range(1, ncol + 1):
            tcell = ws.cell(title_r, c)
            tcell.fill = PatternFill("solid", fgColor=NAVY_D)
            tcell.border = Border(bottom=Side(style="medium", color=ACCENT))
        ws.row_dimensions[title_r].height = 32

    headers = {}
    for c in range(1, ncol + 1):
        cell = ws.cell(hdr_r, c)
        headers[c] = cell.value
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Yu Gothic", size=11, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[hdr_r].height = 30

    for r in range(hdr_r + 1, nrow + 1):
        total = is_total_row(ws, r, hdr_r, ncol)
        band = (r - hdr_r) % 2 == 0
        for c in range(1, ncol + 1):
            cell = ws.cell(r, c)
            h = headers.get(c)
            cell.border = Border(left=thin, right=thin, top=TOP_NAVY, bottom=thin) if total else BORDER
            if total:
                cell.fill = PatternFill("solid", fgColor=TOTAL)
            elif band:
                cell.fill = PatternFill("solid", fgColor=BAND)
            num = is_num_col(h) and isinstance(cell.value, (int, float))
            cell.alignment = Alignment(
                horizontal="right" if num else ("center" if c == 1 else "left"),
                vertical="center", wrap_text=False,
            )
            color = TXT
            v = str(cell.value or "")
            if isinstance(cell.value, (int, float)) and cell.value < 0:
                color = RED
            elif any(k in v for k in ("確定", "入金済", "完了", "計上")):
                color = GREEN
            elif any(k in v for k in ("進行中", "未", "予定")):
                color = AMBER
            cell.font = Font(name="Yu Gothic", size=10, color=color,
                             bold=total or (c == 1 and h == "順位"))
            if num and is_pct_col(h):
                cell.number_format = "0.0%"
            elif num:
                cell.number_format = "#,##0"

    for c in range(1, ncol + 1):
        w = cjk_w(headers.get(c))
        for r in range(hdr_r + 1, min(nrow, hdr_r + 400) + 1):
            w = max(w, cjk_w(ws.cell(r, c).value))
        long_text = any(k in str(headers.get(c) or "") for k in ("顧客", "備考", "説明", "銀行", "売主", "金融", "次回", "内容", "主要"))
        cap = 46 if long_text else 22
        ws.column_dimensions[get_column_letter(c)].width = min(max(w * 0.62 + 3, 7), cap)
        if long_text:
            for r in range(hdr_r + 1, nrow + 1):
                ws.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 条件付き書式: 利益/利益(万) にデータバー、利益率/構成比 にカラースケール
    last_data = nrow
    while last_data > hdr_r and is_total_row(ws, last_data, hdr_r, ncol):
        last_data -= 1   # 合計行はバー対象から外す
    if last_data > hdr_r + 1:
        for c in range(1, ncol + 1):
            h = str(headers.get(c) or "")
            col = get_column_letter(c)
            rng = f"{col}{hdr_r+1}:{col}{last_data}"
            if h in ("利益(万)", "利益") or h.startswith("利益(万"):
                ws.conditional_formatting.add(rng, DataBarRule(
                    start_type="min", end_type="max", color="9DB8D9", showValue=True))
            elif ("利益率" in h) or ("構成比" in h):
                ws.conditional_formatting.add(rng, ColorScaleRule(
                    start_type="min", start_color="FFFFFF",
                    end_type="max", end_color="BFE3C9"))

    # 固定枠: 横長シートは先頭の識別列も固定
    key_cols = 1
    if ncol >= 8:
        key_cols = 3 if (headers.get(1) in ("順位", "年月", "期") or "年月" in str(headers.get(2))) else 1
    ws.freeze_panes = ws.cell(hdr_r + 1, key_cols + 1).coordinate

def main(path):
    wb = openpyxl.load_workbook(path)
    palette = ["1F3A5F", "2E6B4F", "7A4E1E", "5B3A6B", "1E5A6B", "6B1E3A"]
    for i, ws in enumerate(wb.worksheets):
        beautify_sheet(ws)
        ws.sheet_properties.tabColor = palette[i % len(palette)]
    wb.save(path)
    print("styled:", path, "->", wb.sheetnames)

if __name__ == "__main__":
    for p in sys.argv[1:]:
        main(p)
